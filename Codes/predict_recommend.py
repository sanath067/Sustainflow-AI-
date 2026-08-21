import os
import warnings

import joblib
import numpy as np
import pandas as pd

from openpyxl import Workbook, load_workbook
from openpyxl.styles import (
    Font,
    PatternFill,
    Alignment,
    Border,
    Side
)


warnings.filterwarnings("ignore")


# ============================================================
# PATHS
# ============================================================

DATA_PATH = os.path.join(
    "Datasets",
    "Fao_merged_clean.csv"
)

MODEL_PATH = os.path.join(
    "Results",
    "fao_model.joblib"
)

RESULTS_DIR = "Results"

EXCEL_PATH = os.path.join(
    RESULTS_DIR,
    "final_results.xlsx"
)


# ============================================================
# FEATURES AND TARGET
# ============================================================

CATEGORICAL_FEATURES = [
    "commodity",
    "country",
    "food_supply_stage"
]

NUMERIC_FEATURES = [
    "year"
]

TARGET = "loss_percentage"


# ============================================================
# CREATE RESULTS DIRECTORY
# ============================================================

os.makedirs(
    RESULTS_DIR,
    exist_ok=True
)


# ============================================================
# LOAD DATASET
# ============================================================

print("\nLoading FAO dataset...")


if not os.path.exists(DATA_PATH):

    raise FileNotFoundError(
        f"Dataset not found: {DATA_PATH}"
    )


df = pd.read_csv(
    DATA_PATH
)


print(
    f"Dataset loaded successfully. "
    f"Rows: {len(df)}"
)


# ============================================================
# LOAD TRAINED MODEL
# ============================================================

print(
    "Loading trained model..."
)


if not os.path.exists(MODEL_PATH):

    raise FileNotFoundError(
        f"\nTrained model not found: {MODEL_PATH}\n\n"
        "Please train the model first using:\n"
        "python Codes/train_fao_model.py"
    )


model = joblib.load(
    MODEL_PATH
)


print(
    "Model loaded successfully."
)


# ============================================================
# GET AVAILABLE VALUES
# ============================================================

commodities = sorted(
    df["commodity"]
    .dropna()
    .astype(str)
    .unique()
)


countries = sorted(
    df["country"]
    .dropna()
    .astype(str)
    .unique()
)


stages = sorted(
    df["food_supply_stage"]
    .dropna()
    .astype(str)
    .unique()
)


years = sorted(
    df["year"]
    .dropna()
    .unique()
)


# ============================================================
# NUMBERED CHOICE FUNCTION
# ============================================================

def choose_from_list(
    title,
    values
):

    print("\n")

    print("=" * 60)

    print(
        title
    )

    print("=" * 60)


    for index, value in enumerate(
        values,
        start=1
    ):

        print(
            f"{index}. {value}"
        )


    while True:

        try:

            choice = int(
                input(
                    f"\nEnter your choice "
                    f"(1-{len(values)}): "
                )
            )


            if 1 <= choice <= len(values):

                return values[
                    choice - 1
                ]


            print(
                "\nInvalid choice. "
                "Please try again."
            )


        except ValueError:

            print(
                "\nPlease enter a valid number."
            )


# ============================================================
# USER INPUT
# ============================================================

selected_commodity = choose_from_list(
    "SELECT FOOD ITEM",
    commodities
)


selected_country = choose_from_list(
    "SELECT COUNTRY",
    countries
)


selected_stage = choose_from_list(
    "SELECT FOOD SUPPLY STAGE",
    stages
)


selected_year = choose_from_list(
    "SELECT YEAR",
    years
)


# ============================================================
# CREATE MODEL INPUT
# ============================================================

input_data = pd.DataFrame(
    {
        "commodity": [
            selected_commodity
        ],

        "country": [
            selected_country
        ],

        "food_supply_stage": [
            selected_stage
        ],

        "year": [
            selected_year
        ]
    }
)


# ============================================================
# MAKE PREDICTION
# ============================================================

prediction = float(
    model.predict(
        input_data
    )[0]
)


# Keep prediction between 0 and 100

prediction = max(
    0,
    min(
        100,
        prediction
    )
)


# ============================================================
# CHECK FOR EXACT DATASET MATCH
# ============================================================

exact_match = df[
    (
        df["commodity"]
        .astype(str)
        == str(selected_commodity)
    )
    &
    (
        df["country"]
        .astype(str)
        == str(selected_country)
    )
    &
    (
        df["food_supply_stage"]
        .astype(str)
        == str(selected_stage)
    )
    &
    (
        df["year"]
        == selected_year
    )
]


actual_loss = None

prediction_error = None


if len(exact_match) > 0:

    actual_loss = float(
        exact_match[
            TARGET
        ].mean()
    )

    prediction_error = abs(
        actual_loss -
        prediction
    )

    data_status = (
        "Exact combination found in FAO dataset"
    )


else:

    data_status = (
        "Combination not directly found in FAO dataset"
    )


# ============================================================
# GET FEATURE IMPORTANCES
# ============================================================

def get_feature_importances(
    pipeline
):

    preprocessor = pipeline.named_steps[
        "preprocess"
    ]

    trained_model = pipeline.named_steps[
        "model"
    ]


    ohe = preprocessor.named_transformers_[
        "cat"
    ]


    categorical_feature_names = list(
        ohe.get_feature_names_out(
            CATEGORICAL_FEATURES
        )
    )


    all_feature_names = (
        categorical_feature_names +
        NUMERIC_FEATURES
    )


    importances = (
        trained_model.feature_importances_
    )


    return pd.DataFrame(
        {
            "feature": all_feature_names,
            "importance": importances
        }
    )


importance_df = get_feature_importances(
    model
)


# ============================================================
# CALCULATE FEATURE GROUP INFLUENCE
# ============================================================

commodity_importance = importance_df[
    importance_df["feature"]
    .str.startswith(
        "commodity_"
    )
]["importance"].sum()


country_importance = importance_df[
    importance_df["feature"]
    .str.startswith(
        "country_"
    )
]["importance"].sum()


stage_importance = importance_df[
    importance_df["feature"]
    .str.startswith(
        "food_supply_stage_"
    )
]["importance"].sum()


# IMPORTANT:
# This is the corrected line that prevents KeyError: False

year_importance = importance_df[
    importance_df["feature"] == "year"
]["importance"].sum()


total_importance = (
    commodity_importance +
    country_importance +
    stage_importance +
    year_importance
)


if total_importance == 0:

    total_importance = 1


commodity_percent = (
    commodity_importance /
    total_importance
) * 100


country_percent = (
    country_importance /
    total_importance
) * 100


stage_percent = (
    stage_importance /
    total_importance
) * 100


year_percent = (
    year_importance /
    total_importance
) * 100


# ============================================================
# INFLUENCE LEVEL
# ============================================================

def get_influence_level(
    influence
):

    if influence >= 50:

        return "HIGH"

    elif influence >= 20:

        return "MEDIUM"

    else:

        return "LOW"


# ============================================================
# EFFECT DIRECTION
# ============================================================

def get_effect(
    column_name,
    selected_value
):

    overall_mean = df[
        TARGET
    ].mean()


    if column_name == "commodity":

        subset = df[
            df["commodity"]
            .astype(str)
            == str(selected_value)
        ]


    elif column_name == "country":

        subset = df[
            df["country"]
            .astype(str)
            == str(selected_value)
        ]


    elif column_name == "food_supply_stage":

        subset = df[
            df["food_supply_stage"]
            .astype(str)
            == str(selected_value)
        ]


    elif column_name == "year":

        subset = df[
            df["year"]
            == selected_value
        ]


    else:

        subset = df


    selected_mean = subset[
        TARGET
    ].mean()


    if selected_mean >= overall_mean:

        return (
            "Increased predicted food loss"
        )

    else:

        return (
            "Decreased predicted food loss"
        )


# ============================================================
# CREATE XAI RESULTS
# ============================================================

explanations = [

    {
        "Parameter":
        "Commodity",

        "Selected Value":
        selected_commodity,

        "Influence (%)":
        round(
            commodity_percent,
            2
        ),

        "Effect":
        get_effect(
            "commodity",
            selected_commodity
        ),

        "Influence Level":
        get_influence_level(
            commodity_percent
        )
    },


    {
        "Parameter":
        "Country",

        "Selected Value":
        selected_country,

        "Influence (%)":
        round(
            country_percent,
            2
        ),

        "Effect":
        get_effect(
            "country",
            selected_country
        ),

        "Influence Level":
        get_influence_level(
            country_percent
        )
    },


    {
        "Parameter":
        "Food Supply Stage",

        "Selected Value":
        selected_stage,

        "Influence (%)":
        round(
            stage_percent,
            2
        ),

        "Effect":
        get_effect(
            "food_supply_stage",
            selected_stage
        ),

        "Influence Level":
        get_influence_level(
            stage_percent
        )
    },


    {
        "Parameter":
        "Year",

        "Selected Value":
        selected_year,

        "Influence (%)":
        round(
            year_percent,
            2
        ),

        "Effect":
        get_effect(
            "year",
            selected_year
        ),

        "Influence Level":
        get_influence_level(
            year_percent
        )
    }

]


explanation_df = pd.DataFrame(
    explanations
)


# ============================================================
# FIND STRONGEST FACTOR
# ============================================================

strongest_factor = explanation_df.loc[
    explanation_df[
        "Influence (%)"
    ].idxmax()
]


# ============================================================
# DISPLAY PREDICTION
# ============================================================

print("\n")

print("=" * 60)

print(
    "PREDICTION RESULT"
)

print("=" * 60)


print(
    f"\nFood item          : "
    f"{selected_commodity}"
)


print(
    f"Country            : "
    f"{selected_country}"
)


print(
    f"Food supply stage  : "
    f"{selected_stage}"
)


print(
    f"Year               : "
    f"{selected_year}"
)


print(
    f"\nPredicted food loss: "
    f"{prediction:.2f}%"
)


print(
    f"\nData status        : "
    f"{data_status}"
)


if actual_loss is not None:

    print(
        f"\nActual food loss   : "
        f"{actual_loss:.2f}%"
    )

    print(
        f"Prediction error   : "
        f"{prediction_error:.2f} "
        f"percentage points"
    )


else:

    print(
        "\nActual food loss   : "
        "Not available"
    )

    print(
        "Prediction error   : "
        "Not available"
    )

    print(
        "\nThis exact combination is not "
        "present in the dataset."
    )

    print(
        "The displayed loss percentage "
        "is a model-based estimate."
    )


# ============================================================
# DISPLAY EXPLAINABLE AI ANALYSIS
# ============================================================

print("\n")

print("=" * 60)

print(
    "EXPLAINABLE AI ANALYSIS"
)

print("=" * 60)


print(
    f"\nAnalyzing why the model produced "
    f"a prediction of {prediction:.2f}%..."
)


print(
    "\nThe following factors influenced "
    "this prediction:\n"
)


for index, row in explanation_df.iterrows():

    print(
        f"{index + 1}. "
        f"{row['Parameter']}"
    )

    print(
        f"   Selected value : "
        f"{row['Selected Value']}"
    )

    print(
        f"   Effect         : "
        f"{row['Effect']}"
    )

    print(
        f"   Influence      : "
        f"{row['Influence (%)']:.2f}%"
    )

    print(
        f"   Influence level: "
        f"{row['Influence Level']}"
    )

    print()


# ============================================================
# EXPLANATION SUMMARY
# ============================================================

print("-" * 60)

print(
    "EXPLANATION SUMMARY"
)

print("-" * 60)


print(
    f"\nPredicted food loss: "
    f"{prediction:.2f}%"
)


print(
    "\nThe strongest influencing factor was:\n"
)


print(
    strongest_factor[
        "Parameter"
    ]
)


print(
    f"Selected value : "
    f"{strongest_factor['Selected Value']}"
)


print(
    f"Effect         : "
    f"{strongest_factor['Effect']}"
)


print(
    f"Influence      : "
    f"{strongest_factor['Influence (%)']:.2f}%"
)


print(
    f"Influence level: "
    f"{strongest_factor['Influence Level']}"
)


# ============================================================
# EXCEL STYLING
# ============================================================

thin_side = Side(
    style="thin"
)


thin_border = Border(
    left=thin_side,
    right=thin_side,
    top=thin_side,
    bottom=thin_side
)


header_fill = PatternFill(
    fill_type="solid",
    fgColor="1F4E78"
)


section_fill = PatternFill(
    fill_type="solid",
    fgColor="D9EAF7"
)


header_font = Font(
    bold=True,
    color="FFFFFF"
)


bold_font = Font(
    bold=True
)


center_alignment = Alignment(
    horizontal="center",
    vertical="center",
    wrap_text=True
)


left_alignment = Alignment(
    horizontal="left",
    vertical="center",
    wrap_text=True
)


# ============================================================
# OPEN OR CREATE EXCEL FILE SAFELY
# ============================================================

if os.path.exists(
    EXCEL_PATH
):

    workbook = load_workbook(
        EXCEL_PATH
    )


    if "Prediction Results" in workbook.sheetnames:

        worksheet = workbook[
            "Prediction Results"
        ]


    else:

        worksheet = workbook.create_sheet(
            "Prediction Results"
        )


else:

    workbook = Workbook()

    worksheet = workbook.active

    worksheet.title = (
        "Prediction Results"
    )


# ============================================================
# FIND NEXT AVAILABLE ROW
# ============================================================

if (
    worksheet.max_row == 1
    and
    worksheet["A1"].value is None
):

    start_row = 1


else:

    start_row = worksheet.max_row + 3


# ============================================================
# WRITE PREDICTION SUMMARY
# ============================================================

worksheet.merge_cells(
    start_row=start_row,
    start_column=1,
    end_row=start_row,
    end_column=5
)


cell = worksheet.cell(
    start_row,
    1
)


cell.value = (
    "PREDICTION SUMMARY"
)

cell.font = header_font

cell.fill = header_fill

cell.alignment = center_alignment


summary_data = [

    (
        "Food Item",
        selected_commodity
    ),

    (
        "Country",
        selected_country
    ),

    (
        "Food Supply Stage",
        selected_stage
    ),

    (
        "Year",
        selected_year
    ),

    (
        "Predicted Food Loss",
        f"{prediction:.2f}%"
    ),

    (
        "Data Status",
        data_status
    )

]


if actual_loss is not None:

    summary_data.append(
        (
            "Actual Food Loss",
            f"{actual_loss:.2f}%"
        )
    )

    summary_data.append(
        (
            "Prediction Error",
            f"{prediction_error:.2f} percentage points"
        )
    )


else:

    summary_data.append(
        (
            "Actual Food Loss",
            "Not available"
        )
    )

    summary_data.append(
        (
            "Prediction Error",
            "Not available"
        )
    )


row_number = start_row + 1


for label, value in summary_data:

    worksheet.cell(
        row_number,
        1
    ).value = label

    worksheet.cell(
        row_number,
        2
    ).value = value

    worksheet.cell(
        row_number,
        1
    ).font = bold_font

    worksheet.cell(
        row_number,
        1
    ).fill = section_fill

    worksheet.cell(
        row_number,
        1
    ).border = thin_border

    worksheet.cell(
        row_number,
        2
    ).border = thin_border

    worksheet.cell(
        row_number,
        1
    ).alignment = left_alignment

    worksheet.cell(
        row_number,
        2
    ).alignment = left_alignment

    row_number += 1


# ============================================================
# WRITE XAI SECTION
# ============================================================

row_number += 1


worksheet.merge_cells(
    start_row=row_number,
    start_column=1,
    end_row=row_number,
    end_column=5
)


cell = worksheet.cell(
    row_number,
    1
)


cell.value = (
    "EXPLAINABLE AI ANALYSIS"
)

cell.font = header_font

cell.fill = header_fill

cell.alignment = center_alignment


row_number += 1


headers = [

    "Parameter",

    "Selected Value",

    "Influence (%)",

    "Effect",

    "Influence Level"

]


for column, header in enumerate(
    headers,
    start=1
):

    cell = worksheet.cell(
        row_number,
        column
    )

    cell.value = header

    cell.font = bold_font

    cell.fill = section_fill

    cell.border = thin_border

    cell.alignment = center_alignment


row_number += 1


for _, explanation in explanation_df.iterrows():

    values = [

        explanation[
            "Parameter"
        ],

        explanation[
            "Selected Value"
        ],

        f"{explanation['Influence (%)']:.2f}%",

        explanation[
            "Effect"
        ],

        explanation[
            "Influence Level"
        ]

    ]


    for column, value in enumerate(
        values,
        start=1
    ):

        cell = worksheet.cell(
            row_number,
            column
        )

        cell.value = value

        cell.border = thin_border


        if column in [3, 5]:

            cell.alignment = center_alignment

        else:

            cell.alignment = left_alignment


    row_number += 1


# ============================================================
# WRITE EXPLANATION SUMMARY
# ============================================================

row_number += 1


worksheet.merge_cells(
    start_row=row_number,
    start_column=1,
    end_row=row_number,
    end_column=5
)


cell = worksheet.cell(
    row_number,
    1
)


cell.value = (
    "EXPLANATION SUMMARY"
)

cell.font = header_font

cell.fill = header_fill

cell.alignment = center_alignment


row_number += 1


summary_explanation = [

    (
        "Strongest Influencing Factor",
        strongest_factor[
            "Parameter"
        ]
    ),

    (
        "Selected Value",
        strongest_factor[
            "Selected Value"
        ]
    ),

    (
        "Effect",
        strongest_factor[
            "Effect"
        ]
    ),

    (
        "Influence",
        f"{strongest_factor['Influence (%)']:.2f}%"
    ),

    (
        "Influence Level",
        strongest_factor[
            "Influence Level"
        ]
    )

]


for label, value in summary_explanation:

    worksheet.cell(
        row_number,
        1
    ).value = label

    worksheet.cell(
        row_number,
        2
    ).value = value

    worksheet.cell(
        row_number,
        1
    ).font = bold_font

    worksheet.cell(
        row_number,
        1
    ).fill = section_fill

    worksheet.cell(
        row_number,
        1
    ).border = thin_border

    worksheet.cell(
        row_number,
        2
    ).border = thin_border

    worksheet.cell(
        row_number,
        1
    ).alignment = left_alignment

    worksheet.cell(
        row_number,
        2
    ).alignment = left_alignment

    row_number += 1


# ============================================================
# SET COLUMN WIDTHS
# ============================================================

worksheet.column_dimensions[
    "A"
].width = 32


worksheet.column_dimensions[
    "B"
].width = 35


worksheet.column_dimensions[
    "C"
].width = 18


worksheet.column_dimensions[
    "D"
].width = 40


worksheet.column_dimensions[
    "E"
].width = 20


# ============================================================
# SAVE EXCEL FILE
# ============================================================

try:

    workbook.save(
        EXCEL_PATH
    )


except PermissionError:

    print(
        "\nERROR: final_results.xlsx "
        "is currently open."
    )

    print(
        "Close the Excel file and run "
        "the program again."
    )

    raise SystemExit


# ============================================================
# FINAL SUCCESS MESSAGE
# ============================================================

print("\n")

print("=" * 60)

print(
    "PREDICTION AND EXPLAINABLE AI ANALYSIS "
    "COMPLETED SUCCESSFULLY"
)

print("=" * 60)


print(
    f"\nResults saved to:"
)

print(
    EXCEL_PATH
)


print(
    "\nPrevious predictions are preserved."
)


print(
    "The new prediction has been added "
    "below the previous results."
)